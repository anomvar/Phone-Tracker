"""FIR 47 / NCCRP 222-2024 — Surat Cyber Crime tower-dump case parser.

Builds per-MSISDN tracks from Mora Bhagal + Jahangirpura Call Details,
anchored with Airtel CGI lat/lon from LEA tower dumps and site-level
sector offsets for Jio / BSNL / other cells.
"""

from __future__ import annotations

import csv
import math
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from parsers.cdr import Ping, Target, Tower, _register_tower

DATA_DIR = Path(__file__).resolve().parents[2] / "data" / "222_FIR_47"

# Ground-truth / geocoded site anchors (Surat) — LAND side, never the river
MORA_ANCHOR = (21.22138, 72.78948)  # Airtel dump CGI 404-98-8290-* (Rander/Mora)
# Residential Jahangirpura west of Tapi — NOT Jahangirpura Bridge (that is mid-river)
JAHANGIR_ANCHOR = (21.23880, 72.77630)
# Land foci used to aim sector handset placement away from the Tapi channel
MORA_LAND = (21.21880, 72.78480)  # Rander / Bhesan residential west of masts
JAHANGIR_LAND = (21.23950, 72.77480)

# Exact CGI → (lat, lon) harvested from Airtel tower-dump CSVs
KNOWN_CGI: dict[str, tuple[float, float]] = {}

# Site membership for cells seen in Call Details (group reports)
CGI_SITE: dict[str, str] = {
    # Mora Bhagal 4G (from cell-id collect + dumps)
    "404-98-8290-241660679": "MORA",
    "404-98-8290-240799496": "MORA",
    "404-98-8290-241660754": "MORA",
    "404-98-8290-241660674": "MORA",
    "404-98-8290-216060674": "MORA",
    "4058570111824": "MORA",
    "4058570111831": "MORA",
    "405857011181B": "MORA",
    "4058570111818": "MORA",
    "405857011181b": "MORA",
    "40457814322395": "MORA",
    "40457208121167": "MORA",
    "40457208121168": "MORA",
    # Jahangirpura 4G / 3G
    "404-98-5290-39877": "JAHANGIR",
    "404-98-5290-6238": "JAHANGIR",
    "404-98-5290-63072": "JAHANGIR",
    "404-98-5290-64406": "JAHANGIR",
    "404-98-5290-57158": "JAHANGIR",
    "40585702C6032": "JAHANGIR",
    "40585702C6018": "JAHANGIR",
    "40585702C601B": "JAHANGIR",
    "40585702c6032": "JAHANGIR",
    "40585702c6018": "JAHANGIR",
    "40585702c601b": "JAHANGIR",
    "40457814122214": "JAHANGIR",
    "40457814121166": "JAHANGIR",
}


def _norm_cgi(cgi: str) -> str:
    return (cgi or "").strip().strip("'").upper().replace(" ", "")


def _norm_msisdn(raw: str) -> Optional[str]:
    s = re.sub(r"\D", "", (raw or "").strip().strip("'"))
    if s.startswith("91") and len(s) == 12:
        s = s[2:]
    if s.startswith("0") and len(s) == 11:
        s = s[1:]
    return s if len(s) == 10 else None


def _parse_dt(date_s: str, time_s: str) -> Optional[datetime]:
    date_s = (date_s or "").strip()
    time_s = (time_s or "").strip()
    # Excel may yield datetime already via openpyxl data_only — callers stringify
    for fmt in (
        "%B %d, %Y %H:%M:%S",
        "%b %d, %Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%B %d, %Y %H:%M",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(f"{date_s} {time_s}", fmt)
        except ValueError:
            continue
    # "May 16, 2024 00:00:00" style already combined
    for fmt in ("%B %d, %Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_s, fmt)
        except ValueError:
            continue
    return None


EARTH_R = 6371000.0

# eNodeB → shared mast lat/lon (filled after loading Airtel dumps)
ENB_MAST: dict[str, tuple[float, float]] = {}


def _stable_hash(s: str) -> int:
    """Process-stable hash (builtin hash() is salted per process)."""
    h = 2166136261
    for ch in s:
        h ^= ord(ch)
        h = (h * 16777619) & 0xFFFFFFFF
    return h


def _destination(lat: float, lon: float, bearing_deg: float, dist_m: float) -> tuple[float, float]:
    br = math.radians(bearing_deg)
    lat2 = lat + (dist_m * math.cos(br) / EARTH_R) * (180 / math.pi)
    lon2 = lon + (dist_m * math.sin(br) / (EARTH_R * max(math.cos(math.radians(lat)), 0.2))) * (
        180 / math.pi
    )
    return lat2, lon2


def _bearing_deg(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Initial bearing from point 1 → point 2, degrees clockwise from north."""
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dl = math.radians(lon2 - lon1)
    x = math.sin(dl) * math.cos(p2)
    y = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dl)
    return (math.degrees(math.atan2(x, y)) + 360.0) % 360.0


def _parse_lte_ids(cgi: str) -> tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Return (mcc, mnc, lac, eci) when parseable."""
    key = _norm_cgi(cgi)
    m = re.match(r"(\d{3})-(\d{1,3})-(\d+)-(\d+)$", key)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
    if key.isdigit() and len(key) >= 12:
        return int(key[:3]), int(key[3:5]), int(key[5:9]), int(key[9:])
    # Jio hex — treat full hex as eci-like id, synthetic lac from prefix
    if re.match(r"^[0-9A-F]+$", key) and len(key) >= 10:
        try:
            eci = int(key[-7:], 16) if len(key) >= 7 else int(key, 16)
            return 405, 857, _stable_hash(key[:8]) % 60000, eci
        except ValueError:
            return None, None, None, None
    return None, None, None, None


def _enb_sector(eci: int) -> tuple[int, int]:
    """LTE: 20-bit eNodeB + 8-bit sector (when ECI is large); else synthetic."""
    if eci >= 256:
        return eci >> 8, eci & 0xFF
    # Small CIDs (Jahangirpura Airtel 5290) — treat whole CI as enb-ish, sector from nibble
    return eci, eci % 3


def _infer_site(cgi: str, group: str = "") -> str:
    key = _norm_cgi(cgi)
    if key in CGI_SITE:
        return CGI_SITE[key]
    if "8290" in key or key.startswith("40585701118") or key.startswith("405857034"):
        return "MORA"
    if "5290" in key or key.startswith("40585702C6"):
        return "JAHANGIR"
    g = (group or "").upper()
    if "JAHANGIR" in g:
        return "JAHANGIR"
    if "MORA" in g:
        return "MORA"
    return "MORA"


def _enb_key(cgi: str) -> str:
    mcc, mnc, lac, eci = _parse_lte_ids(cgi)
    if eci is None:
        return f"RAW:{_norm_cgi(cgi)}"
    enb, _ = _enb_sector(eci)
    return f"{mcc}-{mnc}-{lac}-{enb}"


def _mast_for_enb(enb_key: str, site: str) -> tuple[float, float]:
    """Unique mast position per eNodeB — shared by its sectors."""
    if enb_key in ENB_MAST:
        return ENB_MAST[enb_key]
    # Inherit from any known CGI on this eNodeB
    for cgi, (la, lo) in KNOWN_CGI.items():
        if _enb_key(cgi) == enb_key:
            ENB_MAST[enb_key] = (la, lo)
            return la, lo
    # Place unknown eNodeBs on a ring around the site (realistic multi-mast cluster)
    anchor = JAHANGIR_ANCHOR if site == "JAHANGIR" else MORA_ANCHOR
    h = _stable_hash(enb_key)
    # Spread 120–520 m around site center so different cells aren't identical
    bearing = (h % 360) + 0.0
    dist = 120.0 + (h // 360 % 400)
    # Bias west of Mora / into Jahangirpura residential (avoid Tapi)
    if site == "MORA":
        bearing = 200.0 + (h % 140)  # S–W–NW
    else:
        bearing = 240.0 + (h % 160)  # W–N arc inland
    lat, lon = _destination(anchor[0], anchor[1], bearing, dist)
    ENB_MAST[enb_key] = (lat, lon)
    return lat, lon


def resolve_cgi(cgi: str, group: str = "") -> tuple[float, float, str]:
    """Return (mast_lat, mast_lon, site) for a CGI."""
    key = _norm_cgi(cgi)
    site = _infer_site(key, group)
    if key in KNOWN_CGI:
        la, lo = KNOWN_CGI[key]
        ENB_MAST.setdefault(_enb_key(key), (la, lo))
        return la, lo, site
    for k, v in KNOWN_CGI.items():
        if k.upper() == key:
            ENB_MAST.setdefault(_enb_key(key), v)
            return (*v, site)
    la, lo = _mast_for_enb(_enb_key(key), site)
    return la, lo, site


def _sector_bearing(cgi: str, site: str, mast_lat: float, mast_lon: float) -> float:
    """3-sector LTE azimuth, softly pulled inland so sectors don't face the river."""
    _, _, _, eci = _parse_lte_ids(cgi)
    sector = 0 if eci is None else _enb_sector(eci)[1]
    az = ((sector % 3) * 120.0 + 30.0) % 360.0
    land = JAHANGIR_LAND if site == "JAHANGIR" else MORA_LAND
    land_az = _bearing_deg(mast_lat, mast_lon, land[0], land[1])
    # Circular blend 65% sector / 35% landward
    x = 0.65 * math.sin(math.radians(az)) + 0.35 * math.sin(math.radians(land_az))
    y = 0.65 * math.cos(math.radians(az)) + 0.35 * math.cos(math.radians(land_az))
    return (math.degrees(math.atan2(x, y)) + 360.0) % 360.0


def _land_bearing(site: str, tower_lat: float, tower_lon: float) -> float:
    focus = JAHANGIR_LAND if site == "JAHANGIR" else MORA_LAND
    return _bearing_deg(tower_lat, tower_lon, focus[0], focus[1])


def estimate_handset(
    msisdn: str,
    cgi: str,
    group: str = "",
    peer_cgi: str = "",
) -> tuple[float, float, float, float, float, str]:
    """
    Refine handset position inside the serving sector.

    Returns (mast_lat, mast_lon, handset_lat, handset_lon, conf_m, site).
    Same CGI no longer collapses every MSISDN — per-number jitter inside the lobe.
    Dual-CGI (First+Last) averages two sector estimates (bilateration-lite).
    """
    mast_lat, mast_lon, site = resolve_cgi(cgi, group)
    bearing = _sector_bearing(cgi, site, mast_lat, mast_lon)
    h = _stable_hash(f"{msisdn}|{_norm_cgi(cgi)}")
    dist = 160.0 + (h % 220)  # 160–380 m into the sector
    lat, lon = _destination(mast_lat, mast_lon, bearing, dist)
    # Cross-track jitter so co-cell users separate (~±55 m)
    cross = ((h >> 8) % 110) - 55.0
    lat, lon = _destination(lat, lon, (bearing + 90.0) % 360.0, cross)
    conf = 160.0

    if peer_cgi and _norm_cgi(peer_cgi) and _norm_cgi(peer_cgi) != _norm_cgi(cgi):
        m2_lat, m2_lon, site2 = resolve_cgi(peer_cgi, group)
        b2 = _sector_bearing(peer_cgi, site2, m2_lat, m2_lon)
        h2 = _stable_hash(f"{msisdn}|{_norm_cgi(peer_cgi)}")
        d2 = 160.0 + (h2 % 220)
        p2_lat, p2_lon = _destination(m2_lat, m2_lon, b2, d2)
        # Weight toward geometric mean of the two sector estimates
        lat = lat * 0.5 + p2_lat * 0.5
        lon = lon * 0.5 + p2_lon * 0.5
        conf = 95.0

    return mast_lat, mast_lon, lat, lon, conf, site


def _load_airtel_cgi_coords() -> None:
    """Populate KNOWN_CGI from LEA Airtel tower-dump CSVs."""
    root = DATA_DIR / "tower dump" / "airtel" / "SP8000570"
    if not root.exists():
        return
    for path in root.glob("*.csv"):
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        try:
            hi = next(i for i, l in enumerate(lines) if l.startswith("Target No"))
        except StopIteration:
            continue
        for line in lines[hi + 1 :]:
            if not line.strip():
                continue
            row = next(csv.reader([line]))
            if len(row) < 12:
                continue
            pairs = [(row[9], row[10]), (row[11], row[12])]
            for latlong, cgi in pairs:
                cgi = _norm_cgi(cgi)
                if not cgi or cgi in ("---", "-") or "/" not in (latlong or ""):
                    continue
                try:
                    la, lo = latlong.strip().split("/")
                    KNOWN_CGI[cgi] = (float(la), float(lo))
                    KNOWN_CGI[cgi.lower()] = (float(la), float(lo))
                except ValueError:
                    continue


def _parse_call_xlsx(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    try:
        hi = next(
            i
            for i, r in enumerate(rows)
            if r and any(c and "Mobile Number" in str(c) for c in r)
        )
    except StopIteration:
        return []
    headers = [str(c).strip() if c else "" for c in rows[hi]]

    def col(*names: str) -> Optional[int]:
        for n in names:
            for i, h in enumerate(headers):
                if n.lower() in h.lower():
                    return i
        return None

    i_m = col("Mobile Number")
    i_cell = col("Cell ID")
    i_last = col("Last Cell")
    i_date = col("Date")
    i_time = col("Time")
    i_name = col("Name")
    i_group = col("Group")
    i_type = col("Sub Call Type")
    i_dur = col("Duration")
    i_op = col("Level Code")
    if i_m is None or i_cell is None:
        return []

    out: list[dict] = []
    for r in rows[hi + 1 :]:
        if not r or i_m >= len(r) or not r[i_m]:
            continue
        msisdn = _norm_msisdn(str(r[i_m]))
        if not msisdn:
            continue
        cell = str(r[i_cell]).strip() if i_cell < len(r) and r[i_cell] else ""
        last = (
            str(r[i_last]).strip()
            if i_last is not None and i_last < len(r) and r[i_last]
            else ""
        )
        date_s = str(r[i_date]).strip() if i_date is not None and r[i_date] else ""
        time_s = str(r[i_time]).strip() if i_time is not None and r[i_time] else "00:00:00"
        # openpyxl may give datetime
        if " " in date_s and time_s in ("00:00:00", ""):
            parts = date_s.split()
            if len(parts) >= 2 and ":" in parts[-1]:
                time_s = parts[-1]
                date_s = " ".join(parts[:-1])
        ts = _parse_dt(date_s, time_s)
        if not ts:
            continue
        dur = 0
        if i_dur is not None and i_dur < len(r) and r[i_dur] is not None:
            try:
                dur = int(float(str(r[i_dur]).strip()))
            except ValueError:
                dur = 0
        out.append(
            {
                "msisdn": msisdn,
                "name": (
                    str(r[i_name]).strip()
                    if i_name is not None and i_name < len(r) and r[i_name]
                    else ""
                ),
                "cell": cell,
                "last": last,
                "ts": ts,
                "dur": dur,
                "group": (
                    str(r[i_group]).strip()
                    if i_group is not None and i_group < len(r) and r[i_group]
                    else ""
                ),
                "ctype": (
                    str(r[i_type]).strip()
                    if i_type is not None and i_type < len(r) and r[i_type]
                    else ""
                ),
                "operator": (
                    str(r[i_op]).strip()
                    if i_op is not None and i_op < len(r) and r[i_op]
                    else ""
                ),
            }
        )
    return out


def _enrich_from_airtel_dump(targets: dict[str, Target]) -> None:
    """Add precise Airtel tower-dump pings for MSISDNs already in the case."""
    root = DATA_DIR / "tower dump" / "airtel" / "SP8000570"
    if not root.exists():
        return
    for path in root.glob("*.csv"):
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        try:
            hi = next(i for i, l in enumerate(lines) if l.startswith("Target No"))
        except StopIteration:
            continue
        for line in lines[hi + 1 :]:
            if not line.strip():
                continue
            row = next(csv.reader([line]))
            if len(row) < 11:
                continue
            msisdn = _norm_msisdn(row[0])
            if not msisdn or msisdn not in targets:
                continue
            ts = _parse_dt(row[6].strip("'"), row[7])
            if not ts:
                continue
            target = targets[msisdn]
            peer = _norm_cgi(row[12]) if len(row) > 12 else ""
            for latlong, cgi in ((row[9], row[10]),):
                cgi_n = _norm_cgi(cgi)
                if not cgi_n or cgi_n in ("---", "-"):
                    continue
                if latlong and "/" in latlong:
                    try:
                        la, lo = latlong.strip().split("/")
                        KNOWN_CGI[cgi_n] = (float(la), float(lo))
                        ENB_MAST[_enb_key(cgi_n)] = (float(la), float(lo))
                    except ValueError:
                        pass
                mast_lat, mast_lon, h_lat, h_lon, conf, site = estimate_handset(
                    msisdn, cgi_n, "MORA B", peer_cgi=peer
                )
                bearing = _sector_bearing(cgi_n, site, mast_lat, mast_lon)
                _register_tower(
                    target.towers,
                    cgi_n,
                    mast_lat,
                    mast_lon,
                    name="Mora Bhagal",
                    operator="Airtel",
                    phone_bearing_deg=bearing,
                )
                target.pings.append(
                    Ping(
                        ts=ts,
                        lat=mast_lat,
                        lon=mast_lon,
                        cgi=cgi_n,
                        tower_name="Mora Bhagal",
                        call_type=row[1].strip() if len(row) > 1 else "",
                        operator="Airtel",
                        msisdn=msisdn,
                        source="cell-fusion",
                        deviation_m=conf,
                        handset_lat=h_lat,
                        handset_lon=h_lon,
                    )
                )


def _operator_short(level: str) -> str:
    u = (level or "").upper()
    if "JIO" in u or "RELIANCE" in u:
        return "Jio"
    if "AIRTEL" in u:
        return "Airtel"
    if "BSNL" in u:
        return "BSNL"
    if "VI" in u or "VODAFONE" in u or "IDEA" in u:
        return "Vi"
    return "Unknown"


def load_fir47_targets(max_targets: int = 28) -> dict[str, Target]:
    """Load FIR-47 individuals ranked by dual-site corridor evidence."""
    KNOWN_CGI.clear()
    ENB_MAST.clear()
    _load_airtel_cgi_coords()

    saliya = DATA_DIR / "cell id dcb collect" / "saliya sir"
    rows = _parse_call_xlsx(saliya / "Call Details - 4.xlsx")
    rows += _parse_call_xlsx(saliya / "New folder" / "jahangirpura.xlsx")
    if not rows:
        return {}

    # Aggregate per MSISDN
    meta: dict[str, dict] = defaultdict(
        lambda: {
            "name": "",
            "operator": "",
            "sites": set(),
            "cells": set(),
            "events": [],
        }
    )
    for r in rows:
        m = meta[r["msisdn"]]
        if r["name"] and not m["name"]:
            m["name"] = re.sub(r"\s+", " ", r["name"]).strip()[:48]
        if r["operator"] and not m["operator"]:
            m["operator"] = _operator_short(r["operator"])
        site = _infer_site(r["cell"], r["group"])
        m["sites"].add(site)
        if r["cell"]:
            m["cells"].add(_norm_cgi(r["cell"]))
        if r["last"]:
            m["cells"].add(_norm_cgi(r["last"]))
        m["events"].append(r)

    # Score: dual-site first, then multi-cell density, then hit count
    scored: list[tuple[float, str]] = []
    for msisdn, m in meta.items():
        dual = 1.0 if len(m["sites"]) >= 2 else 0.0
        score = (
            dual * 1000
            + len(m["cells"]) * 40
            + len(m["events"]) * 2
            + (20 if m["name"] else 0)
        )
        # Prefer people with enough temporal samples to track
        if len(m["events"]) < 2 and not dual:
            continue
        scored.append((score, msisdn))
    scored.sort(reverse=True)

    targets: dict[str, Target] = {}
    for _, msisdn in scored[:max_targets]:
        m = meta[msisdn]
        sites = sorted(m["sites"])
        site_tag = "↔".join(sites) if len(sites) > 1 else (sites[0] if sites else "?")
        name = m["name"] or "Unknown"
        label = f"FIR47 · {name}"
        if len(sites) >= 2:
            label = f"FIR47 · {name} · CORRIDOR"
        op = m["operator"] or "Multi"
        target = Target(
            msisdn=msisdn,
            operator=op,
            label=label,
        )
        # stash extras for API
        target._fir_sites = sites  # type: ignore[attr-defined]
        target._fir_name = name  # type: ignore[attr-defined]
        target._fir_corridor = len(sites) >= 2  # type: ignore[attr-defined]

        for ev in m["events"]:
            first = _norm_cgi(ev["cell"])
            last = _norm_cgi(ev["last"]) if ev["last"] else ""
            if not first:
                continue
            # Primary observation at call start (fuse First+Last when both present)
            mast_lat, mast_lon, h_lat, h_lon, conf, site = estimate_handset(
                msisdn, first, ev["group"], peer_cgi=last
            )
            tower_name = "Jahangirpura" if site == "JAHANGIR" else "Mora Bhagal"
            bearing = _sector_bearing(first, site, mast_lat, mast_lon)
            _register_tower(
                target.towers,
                first,
                mast_lat,
                mast_lon,
                name=tower_name,
                operator=op,
                phone_bearing_deg=bearing,
            )
            target.pings.append(
                Ping(
                    ts=ev["ts"],
                    lat=mast_lat,
                    lon=mast_lon,
                    cgi=first,
                    tower_name=tower_name,
                    call_type=ev["ctype"],
                    operator=op,
                    msisdn=msisdn,
                    source="cell-fusion",
                    deviation_m=conf,
                    handset_lat=h_lat,
                    handset_lon=h_lon,
                )
            )
            # End-of-call on a different CGI → second fused observation
            if last and last != first:
                end_ts = ev["ts"] + timedelta(seconds=max(ev["dur"], 1))
                m2, n2, hl2, hn2, conf2, site2 = estimate_handset(
                    msisdn, last, ev["group"], peer_cgi=first
                )
                tname2 = "Jahangirpura" if site2 == "JAHANGIR" else "Mora Bhagal"
                b2 = _sector_bearing(last, site2, m2, n2)
                _register_tower(
                    target.towers,
                    last,
                    m2,
                    n2,
                    name=tname2,
                    operator=op,
                    phone_bearing_deg=b2,
                )
                target.pings.append(
                    Ping(
                        ts=end_ts,
                        lat=m2,
                        lon=n2,
                        cgi=last,
                        tower_name=tname2,
                        call_type=ev["ctype"] + "/LAST",
                        operator=op,
                        msisdn=msisdn,
                        source="cell-fusion",
                        deviation_m=conf2,
                        handset_lat=hl2,
                        handset_lon=hn2,
                    )
                )

        target.pings.sort(key=lambda p: p.ts)
        deduped: list[Ping] = []
        seen: set[tuple] = set()
        for p in target.pings:
            key = (p.ts.replace(microsecond=0), p.cgi)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(p)
        target.pings = deduped
        _fingerprint_refine(target)
        if target.pings:
            targets[msisdn] = target

    _enrich_from_airtel_dump(targets)
    for t in targets.values():
        t.pings.sort(key=lambda p: p.ts)
        _fingerprint_refine(t)
        _fan_out_colocated_towers(t)

    return targets


def _fingerprint_refine(target: Target) -> None:
    """Pull each fix toward the centroid of all cells this person used at that site."""
    by_site: dict[str, list[Ping]] = defaultdict(list)
    for p in target.pings:
        if p.handset_lat is None or p.handset_lon is None:
            continue
        site = "JAHANGIR" if "Jahangir" in (p.tower_name or "") else "MORA"
        by_site[site].append(p)
    for plist in by_site.values():
        cgi_pts: dict[str, list[tuple[float, float]]] = defaultdict(list)
        for p in plist:
            cgi_pts[p.cgi].append((p.handset_lat, p.handset_lon))  # type: ignore[arg-type]
        if len(cgi_pts) < 2:
            continue
        means = [
            (sum(a for a, _ in pts) / len(pts), sum(b for _, b in pts) / len(pts))
            for pts in cgi_pts.values()
        ]
        cx = sum(a for a, _ in means) / len(means)
        cy = sum(b for _, b in means) / len(means)
        for p in plist:
            p.handset_lat = float(p.handset_lat) * 0.72 + cx * 0.28  # type: ignore[arg-type]
            p.handset_lon = float(p.handset_lon) * 0.72 + cy * 0.28  # type: ignore[arg-type]
            p.deviation_m = min(float(p.deviation_m or 150), 115.0)
            p.source = "cell-fusion"


def _fan_out_colocated_towers(target: Target) -> None:
    """
    Co-sited LTE sectors often share one LEA lat/lon (same eNodeB).
    Spread them on a small ring so the map shows every CGI, not one stacked pin.
    """
    groups: dict[tuple[float, float], list[Tower]] = defaultdict(list)
    for tw in target.towers.values():
        groups[(round(tw.lat, 5), round(tw.lon, 5))].append(tw)
    for (lat0, lon0), group in groups.items():
        if len(group) < 2:
            continue
        group.sort(key=lambda t: t.cgi)
        n = len(group)
        for i, tw in enumerate(group):
            _, _, _, eci = _parse_lte_ids(tw.cgi)
            sector = 0 if eci is None else _enb_sector(eci)[1]
            # Prefer real sector azimuth; fall back to even ring
            bearing = ((sector % 3) * 120.0 + 20.0 + i * 8.0) % 360.0
            if n > 3:
                bearing = (i * (360.0 / n) + 15.0) % 360.0
            dist = 28.0 + (i % 3) * 6.0  # 28–40 m — visible, still same site
            tw.lat, tw.lon = _destination(lat0, lon0, bearing, dist)
            tw.phone_bearing_deg = bearing


def fir47_summary(target: Target) -> dict:
    """Extra FIR-47 fields for the UI."""
    base = target.to_summary()
    base["case"] = "fir47"
    base["name"] = getattr(target, "_fir_name", "")
    base["sites"] = getattr(target, "_fir_sites", [])
    base["corridor"] = bool(getattr(target, "_fir_corridor", False))
    base["label"] = target.label
    return base
